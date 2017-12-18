#Imports
import sys
import os
import _pickle as cPickle
from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QWidget, QAction, QTabWidget, QVBoxLayout, QDialog,  QFileDialog, QMessageBox, QTableWidgetItem
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer 
from openpyxl import *
from main import Ui_MainWindow

#Ventana principal, la columna del programa
class ActualizadorPrecios(QMainWindow, Ui_MainWindow): 
	def __init__(self,*args,**kwargs):
		QMainWindow.__init__(self,*args,**kwargs)
		self.setupUi(self)

		#Defino un Diccionario que va a guardar la informacion
		#de Cada proveedor.
		self.dicProve = {
		"Royaltek":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"Fispa":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"RM":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"Ferman":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"Crifa":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"DER":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"Electrodisiel":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"DZE":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Titulo":0
		},
		"Expoyer":{
		"Ganancia":0,
		"IVA":0,
		"Descuento":0,
		"Control":0,
		"Precio":0,
		"Hoja":0,
		"Hoja":0,
		"Titulo":0
		}
		}

		#Defino variables
		self.archivo_lista = 0 #Objeto con la plantilla
		self.path = os.getcwd()
		self.prove = 0 #Direccion archivo Proveedor
		self.lista = 0 #Direccion archivo Plantilla ML
		self.hojaProve = 0 #Objeto Hoja XLSX Proveedor
		self.hojaLista =0 #Objeto Hoja XLSX Plantilla

		#Conecto los Botones
		self.buscarEntrada.clicked.connect(self.BuscarEntrada)
		self.buscarSalida.clicked.connect(self.BuscarSalida)
		self.foward_1.clicked.connect(self.Ingreso)
		self.foward_2.clicked.connect(self.Exportar)
		self.toolButton.clicked.connect(self.Opciones)
		self.btnAceptar.clicked.connect(self.Inicio)
		self.comboBox03.activated.connect(self.CompletoOpciones)
		self.btnGuardar.clicked.connect(self.Guardar)
		#Inicio de Funciones
		self.BaseImportar()

		#-- Funciones --
	def Inicio(self):
		self.folio.setCurrentIndex(0)
		self.BaseExportar()
	#Busco el archivo de Entrada con los precios nuevos
	def BuscarEntrada(self):
		self.prove = QFileDialog.getOpenFileName(self, 'Open file',self.path,"(*xlsx)")
		self.entrada.setText(str(self.prove[0]))
		if len(self.prove[0]) < 5:
			QMessageBox.about(self, "Error archivo", "Elija una lista de precios correcta")
			self.foward_1.setEnabled(False)
		else:
			self.foward_1.setEnabled(True)

	#Busco el Archivo de Salida para modificar precios
	def BuscarSalida(self):
		self.lista = QFileDialog.getOpenFileName(self, 'Open file',self.path,"(*xlsx)")
		self.salida.setText(str(self.lista[0]))
		if len(self.lista[0]) < 5:
			QMessageBox.about(self, "Error archivo", "Elija una lista de precios correcta")
			self.foward_1.setEnabled(False)
		else:
			self.foward_1.setEnabled(True)

	#Cargo los Excels a dos Objetos: Proveedor y Lista(plantila ML)
	def CargarExcels(self):
		archivo_proveedor = load_workbook(self.prove[0])
		#Ahora Pido los nombres de las worksheets
		aux = archivo_proveedor.sheetnames
		aux2 = self.dicProve[self.comboBox.currentText()]["Hoja"]

		self.hojaProve = archivo_proveedor[aux[aux2]]
		self.archivo_lista = load_workbook(self.lista[0])
		self.hojaLista = self.archivo_lista.active

	#Guardo el objeto de la plantilla a otro nombre.
	def Exportar(self):
		for x in range(self.tabla.rowCount()):
			celda = self.hojaLista.cell(row=x+2,column=6)
			valorAux = self.tabla.item(x,3).text()
			celda.value = int(valorAux)
		self.archivo_lista.save("Actualizado.xlsx")
		#Termino, cambio a la ultima pantalla
		self.folio.setCurrentIndex(2)

	#Completo las opciones con sus datos
	def CompletoOpciones(self):
		try:
			self.edit01.setText(self.dicProve[self.comboBox03.currentText()]["Ganancia"])
		except:
			self.dicProve[self.comboBox03.currentText()]["Ganancia"] = self.edit01.setText("Complete")
		try:
			self.edit02.setText(self.dicProve[self.comboBox03.currentText()]["IVA"])
		except:
			self.dicProve[self.comboBox03.currentText()]["IVA"] = self.edit02.setText("Complete")
		try:
			self.edit03.setText(self.dicProve[self.comboBox03.currentText()]["Descuento"])
		except:
			self.dicProve[self.comboBox03.currentText()]["Descuento"] = self.edit03.setText("Complete")
		try:
			self.spinBox01.setValue(self.dicProve[self.comboBox03.currentText()]["Control"])
		except:
			self.dicProve[self.comboBox03.currentText()]["Control"] = self.spinBox01.setValue(0)
		try:
			self.spinBox02.setValue(self.dicProve[self.comboBox03.currentText()]["Precio"])
		except:
			self.dicProve[self.comboBox03.currentText()]["Precio"] = self.spinBox02.setValue(0)
		try:
			self.spinBox03.setValue(self.dicProve[self.comboBox03.currentText()]["Hoja"])
		except:
			self.dicProve[self.comboBox03.currentText()]["Hoja"] = self.spinBox03.setValue(0)
		try:
			self.spinBox04.setValue(self.dicProve[self.comboBox03.currentText()]["Titulo"])
		except:
			self.dicProve[self.comboBox03.currentText()]["Titulo"] = self.spinBox04.setValue(0)
		try:
			self.label05.setText(self.edit03.text()+"*"+self.edit02.text()+"*"+self.edit01.text())
		except:
			self.label05.setText("Ingrese datos para la formula")

	def Opciones(self):
		self.folio.setCurrentIndex(3)
		self.CompletoOpciones()
	#Guardo las opciones	
	def Guardar(self):
		self.dicProve[self.comboBox03.currentText()]["Ganancia"] = self.edit01.text()
		self.dicProve[self.comboBox03.currentText()]["IVA"] = self.edit02.text()
		self.dicProve[self.comboBox03.currentText()]["Descuento"] = self.edit03.text()
		self.dicProve[self.comboBox03.currentText()]["Control"] = self.spinBox01.value()
		self.dicProve[self.comboBox03.currentText()]["Precio"] = self.spinBox02.value()
		self.dicProve[self.comboBox03.currentText()]["Hoja"] = self.spinBox03.value()
		self.dicProve[self.comboBox03.currentText()]["Titulo"] = self.spinBox04.value()
		
	
	def BaseExportar(self):
		wb = Workbook()
		ws = wb.active
		llaves = self.dicProve.keys()

		for x,key in enumerate(llaves):
			col = 1
			for val in self.dicProve[key].values():
				celda = ws.cell(row=x+1,column=col)
				celda.value = val
				col= col+1
		wb.save("Base.xlsx")

	def BaseImportar(self):
		"Ahora importo"
		wb = load_workbook('Base.xlsx')
		ws = wb.active
		llaves = self.dicProve.keys()
		
		for x,key in enumerate(llaves):
			col = 1
			aux = self.dicProve[key].items()
			for key2,value in aux:
				celda = ws.cell(row=x+1,column=col)
				self.dicProve[key][key2] = celda.value
				col= col+1

#Aca lo que tengo que hacer es Leer el archivo en Excel y guardar en un Diccionario
#Cada una de las columnas
	def Ingreso(self):
		#Variables Auxiliares
		dicML = {
		"SKU":1,
		"Precio":6,
		"Titulo":2
		}

		SKU = []
		Precio_Viejo = []
		Titulo = []
		rawControl = []
		rawPrecio_Nuevo = []
		listaAuxML=[]
		listaAuxProve=[]
		
		#Abro los Archivos Excel
		self.CargarExcels()
		self.tabla.setRowCount(0)

		#Compruebo que la ultima celda no este vacia.
		filas = self.hojaLista.max_row
		for x in range(self.hojaLista.max_row):
			c = self.hojaLista.cell(row=filas, column=1)
			if any(c.value):
				break
			else:
				filas = filas-1

		for x in range((filas)-1):
			#Hago una copia del SKU, editado para que coincida siempre
			#Esta copia la guardo en memoria y la voy a usar para chequear
			#el codigo con el proveedor. Despues utilizo su indice para saber
			#donde remplazar-
			auxML = self.hojaLista.cell(row=x+2,column=dicML["SKU"])
			auxML = auxML.value
			#-----↕
			SKU.append(self.hojaLista.cell(row=x+2,column=dicML["SKU"]))
			#-----↕
			
			auxML = str(auxML)
			auxML = auxML.upper()
			auxML = auxML.strip("*")
			auxML = auxML.strip()
			listaAuxML.append(auxML)
			Precio_Viejo.append(self.hojaLista.cell(row=x+2,column=dicML["Precio"]))
			Titulo.append(self.hojaLista.cell(row=x+2,column=dicML["Titulo"]))
		
		#-----↕
		x = 0
		#-----↕
		for x in range(self.hojaProve.max_row):

			print(self.dicProve[self.comboBox.currentText()]["Control"])

			auxProve = self.hojaProve.cell(row=x+1,column=self.dicProve[self.comboBox.currentText()]["Control"])
			auxProve = auxProve.value
			#-----↕
			rawControl.append(self.hojaProve.cell(row=x+1,column=self.dicProve[self.comboBox.currentText()]["Control"]))
			#-----↕
			auxProve = str(auxProve)
			auxProve = auxProve.upper()
			auxProve = auxProve.strip("*")
			auxProve = auxProve.strip()
			#print(auxProve)
			listaAuxProve.append(auxProve)
			rawPrecio_Nuevo.append(self.hojaProve.cell(row=x+1,column=self.dicProve[self.comboBox.currentText()]["Precio"]))

		#Completo la tabla con los datos del Excel. *Y* es la fila. 
		for y,sku in enumerate(SKU):

			#Inserto una fila
			self.tabla.insertRow(y)
			self.tabla.setItem(y, 0, QTableWidgetItem(sku.value))
			self.tabla.setItem(y, 2, QTableWidgetItem(str(Precio_Viejo[y].value)))
			self.tabla.setItem(y, 4, QTableWidgetItem(Titulo[y].value))
			#Hago un While para recorrer el SKU del provedor, el AuxProve
			
			#-----↕
			x = 0
			#-----↕

			AuxDes= self.dicProve[self.comboBox.currentText()]["Descuento"]
			AuxIVA= self.dicProve[self.comboBox.currentText()]["IVA"]
			AuxGan= self.dicProve[self.comboBox.currentText()]["Ganancia"]

			#-----↕	
			while x<len(listaAuxProve):
				#print("Comparo "+listaAuxProve[x]+" Con "+listaAuxML[y])
				if listaAuxProve[x] == listaAuxML[y]:
					
					self.tabla.setItem(y,1,QTableWidgetItem(str(rawControl[x].value)))
					money = float(rawPrecio_Nuevo[x].value)
					money= money*float(AuxDes)*float(AuxIVA)*float(AuxGan)
					money = int(round(money))
					self.tabla.setItem(y,3,QTableWidgetItem(str(money)))
					break

				x=x+1

		self.folio.setCurrentIndex(1)

if __name__ == '__main__':
	app = QApplication(sys.argv)
	prog =  ActualizadorPrecios()
	prog.show()
	sys.exit(app.exec_())