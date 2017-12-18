import sys
import os
import _pickle as cPickle
from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QWidget, QAction, QTabWidget, QVBoxLayout, QDialog,  QFileDialog, QMessageBox, QTableWidgetItem
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer 
from openpyxl import *
from main import Ui_MainWindow

#Ventana principal, la columna del programa
class ActualizadorPrecios(QMainWindow, Ui_MainWindow): 
	def __init__(self,*args,**kwargs):
		QMainWindow.__init__(self,*args,**kwargs)
		self.setupUi(self)
		#Defino diccionario con los valores de los proveedores
		dicGanancia = {
		"Ganancia": 1.37,
		"IVA": 1.21,
		"Descuento": 0.55
		}
		dicProve = {

		"RoyalTek":dicGanancia,
		"Crifa":dicGanancia,
		"Fispa":dicGanancia,
		"Ferman":dicGanancia

		}

		#Defino variables
		self.archivo_lista = 0
		self.path = os.getcwd()
		self.prove = 0
		self.lista = 0
		self.hojaProve = 0
		self.hojaLista =0


		#Conecto los Botones
		self.buscarEntrada.clicked.connect(self.BuscarEntrada)
		self.buscarSalida.clicked.connect(self.BuscarSalida)
		self.foward_1.clicked.connect(self.Ingreso)
		self.foward_2.clicked.connect(self.Exportar)

		#-- Funciones --

#Busco el archivo de Entrada con los precios nuevos
	def BuscarEntrada(self):
		self.prove = QFileDialog.getOpenFileName(self, 'Open file',self.path,"(*xlsx)")
		self.entrada.setText(str(self.prove[0]))
		if len(self.prove[0]) < 5:

			QMessageBox.about(self, "Error archivo", "Elija una lista de precios correcta")
			self.foward_1.setEnabled(False)
		else:
			self.foward_1.setEnabled(True)



	#Busco el Archivo de Salida para modificar precios
	def BuscarSalida(self):
		self.lista = QFileDialog.getOpenFileName(self, 'Open file',self.path,"(*xlsx)")
		self.salida.setText(str(self.lista[0]))
		if len(self.lista[0]) < 5:
			QMessageBox.about(self, "Error archivo", "Elija una lista de precios correcta")
			self.foward_1.setEnabled(False)
		else:
			self.foward_1.setEnabled(True)
#Aca lo que tengo que hacer es Leer el archivo en Excel y guardar en un Diccionario
#Cada una de las columnas
	def Ingreso(self):
		dicML = {
		"SKU":1,
		"Precio":6,
		"Titulo":2
		}
		dicProveedor = {
		"Control":3,
		"Precio":5,
		"Descuento":0.55,
		"IVA":1.21,
		"Ganancia":1.37
		}
		SKU = []
		Precio_Viejo = []
		Titulo = []
		rawControl = []
		rawPrecio_Nuevo = []
		listaAuxML=[]
		listaAuxProve=[]

		#Abro los Archivos Excel
		self.CargarExcels()

		#Obtengo datos de mi lista
		self.tabla.setRowCount(0)
		#Compruebo que la ultima celda no este vacia.
		filas = self.hojaLista.max_row
		for x in range(self.hojaLista.max_row):
			c = self.hojaLista.cell(row=filas, column=1)
			if any(c.value):
				break
			else:
				filas = filas-1




		for x in range((filas)-1):
			#Hago una copia del SKU, editado para que coincida siempre
			#Esta copia la guardo en memoria y la voy a usar para chequear
			#el codigo con el proveedor. Despues utilizo su indice para saber
			#donde remplazar-
			auxML = self.hojaLista.cell(row=x+2,column=dicML["SKU"])
			auxML = auxML.value
			#-----↕
			SKU.append(self.hojaLista.cell(row=x+2,column=dicML["SKU"]))
			#-----↕
			
			auxML = str(auxML)
			auxML = auxML.upper()
			auxML = auxML.strip("*")
			auxML = auxML.strip()
			listaAuxML.append(auxML)
			Precio_Viejo.append(self.hojaLista.cell(row=x+2,column=dicML["Precio"]))
			Titulo.append(self.hojaLista.cell(row=x+2,column=dicML["Titulo"]))

		#-----↕
		x = 0
		#-----↕
		for x in range(self.hojaProve.max_row):

			auxProve = self.hojaProve.cell(row=x+1,column=dicProveedor["Control"])
			auxProve = auxProve.value
			#-----↕
			rawControl.append(self.hojaProve.cell(row=x+1,column=dicProveedor["Control"]))
			#-----↕
			auxProve = str(auxProve)
			auxProve = auxProve.upper()
			auxProve = auxProve.strip("*")
			auxProve = auxProve.strip()
			listaAuxProve.append(auxProve)
			rawPrecio_Nuevo.append(self.hojaProve.cell(row=x+1,column=dicProveedor["Precio"]))

	

		#Completo la tabla con los datos del Excel. *Y* es la fila. 
		for y,sku in enumerate(SKU):

			#Inserto una fila
			self.tabla.insertRow(y)
			self.tabla.setItem(y, 0, QTableWidgetItem(sku.value))
			self.tabla.setItem(y, 2, QTableWidgetItem(str(Precio_Viejo[y].value)))
			self.tabla.setItem(y, 4, QTableWidgetItem(Titulo[y].value))
			#Hago un While para recorrer el SKU del provedor, el AuxProve
			
			#-----↕
			x = 0
			#-----↕	
			while x<len(listaAuxProve):
				#print("Comparo "+listaAuxProve[x]+" Con "+listaAuxML[y])
				if listaAuxProve[x] == listaAuxML[y]:
					
					self.tabla.setItem(y,1,QTableWidgetItem(str(rawControl[x].value)))
					money = rawPrecio_Nuevo[x].value*dicProveedor["Descuento"]*dicProveedor["IVA"]*dicProveedor["Ganancia"]
					money = int(round(money))
					self.tabla.setItem(y,3,QTableWidgetItem(str(money)))
					break

				x=x+1

		self.folio.setCurrentIndex(1)	

	def CargarExcels(self):
		archivo_proveedor = load_workbook(self.prove[0])
		self.hojaProve = archivo_proveedor.active
		self.archivo_lista = load_workbook(self.lista[0])
		self.hojaLista = self.archivo_lista.active
		

	def Exportar(self):
		
		for x in range(self.tabla.rowCount()):
			celda = self.hojaLista.cell(row=x+2,column=6)
			valorAux = self.tabla.item(x,3).text()
			celda.value = int(valorAux)
		self.archivo_lista.save("Actualizado.xlsx")

		self.folio.setCurrentIndex(2)
		





if __name__ == '__main__':
	app = QApplication(sys.argv)
	prog =  ActualizadorPrecios()
	prog.show()
	sys.exit(app.exec_())